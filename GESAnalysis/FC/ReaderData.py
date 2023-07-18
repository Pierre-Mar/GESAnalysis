import os
import platform
import csv
from typing import Union, Dict, List, Tuple


class ReaderData:
    """ Class to read data from a CSV, TSV, TXT or XLSX file
    
        Exemple : File format :
            nom_col1,nom_col2.unite,nom_col3.suite_nom.unite,nom_col4.suite_nom.unite.suite_unite
            d11,d12,d13,d14,
            d21,d22,d23,d24
                
            The dictionary will be :
            {
                "nom_col1": { "name": ["nom_col1"], "unit": None, "data": [[d11], [d21]], "type": type(data)},
                "nom_col2.unite": { "name": ["nom_col2"], "unit": ["unite"], "data": [[d12], [d22]] },
                "nom_col3.suite_nom.unite": { "name": ["nom_col3", "suite_nom"], "unit": ["unite"], "data": [[d13], [d23]] },
                "nom_col4.suite_nom.unite.suite_unite": { "name": ["nom_col4", "suite_nom"], "unit": ["unite", "suite_unite"], "data": [[d14], [d24]] },
            }
    """
    
    __accepted_extension = [".csv", ".txt", ".tsv", ".xlsx"]
    
    
    # Units accepted for data
    __units = ["t", "kg", "g", "mg",                    # Units for mass
              "km", "hm", "dam", "m", "dm", "cm", "mm", # Units for length
              "co2e"                                    # Unit for CO2
            ]
    
    
    def __init__(self) -> None:
        """ Initialisation of the class
        """
        pass


    def read_file(self, filename: str, sep: str = None, engine: str ='pandas') -> Dict[str, Dict[str, List[Union[List[Union[int, float, bool, str]], str]]]]:
        """ Read the file 'filename' with or without a separator, for CSV, TSV and TXT files, and
            a reading engine, for XLSX files

        Args:
            filename (str): Path to file
            sep (str, optional): Separator between values in a CSV, TSV and TXT files. Defaults to None
            engine (str, optional): Reading engine for XLSX file (pandas, openpyxl). Defaults to pandas

        Returns:
            dict: Dictionary with the name, unit, data and type of each column in the file if the reading is correct
        """
        # Verification of the file
        self.__verify_file(filename)
        
        # Reading of the CSV, TSV or TXT file with a separator
        if self.__ext in [".csv", ".tsv", ".txt"]:
            # If it's a TSV file, then the separator is '\t'
            if self.__ext == ".tsv":
                delimiter = "\t"
            delimiter = sep
            # If the separator is not given, then it automatically detects
            if delimiter is None:
                delimiter = self.__detect_delimiter(filename)
            return self.__read_csv_tsv_txt(filename, delimiter)
        
        # If it's not a CSV, TSV or TXT file, then it's a XLSX file
        return self.__read_xlsx(filename, engine)
        
    
    def __verify_file(self, filename: str) -> None:        
        """ Check if the file 'filename' exists and it can be read

        Args:
            filename (str): Path to file

        Raises:
            FileNotFoundError: 'filename' was not found
            TypeError: 'filename' is not a CSV, TSV, TXT or XLSX file
        """
        root_filename, self.__ext = os.path.splitext(filename)
        
        # Get the name of file and remove the path
        # The path is different between OS
        os_name = platform.system()
        sep_path = '/'
        if os_name == "Windows":
            sep_path = '\\'
        path_to_file = root_filename.split(sep_path)
        file = path_to_file[len(path_to_file) - 1]
        
        # Check if the file exists
        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Le fichier '{file+self.__ext}' n'existe pas")
        
        # Check if the file is supported by the application
        if not self.__ext in self.__accepted_extension:
            raise TypeError(f"Exportation impossible de '{file+self.__ext}'. Le fichier doit être de type CSV, TSV, TXT ou XLSX")
    
    
    def __detect_delimiter(self, filename: str) -> str:
        """ Detects the separator inside the file 'filename'

        Args:
            filename (str): Path to file

        Returns:
            str: The separator inside the file
        """
        sniffer = csv.Sniffer()
        
        # Get the 1st line (column names) to found the separator
        head_line = None
        with open(filename, "r") as f:
            head_line = f.readline()
        
        dialect = sniffer.sniff(head_line)
        return dialect.delimiter
    

    def __read_csv_tsv_txt(self, filename: str, sep:str = ',') -> Dict[str, Dict[str, List[Union[List[Union[int, float, bool, str]], str]]]]:
        """ Reading the data from a CSV, TSV or TXT file and the separator between values

        Args:
            filename (str): Path to file
            sep (str, optional): Separator between values in the file. Defaults to ','

        Raises:
            ValueError: Number of elements in a row is different from the number of columns
            TypeError: the type of an element is different from the element of his column

        Returns:
            dict: Dictionary with the data of file if the reading is correct
        """
        data = {}
        with open(filename, "r") as file:
            # Read the 1st line (column names)
            column = file.readline()
            
            # Remove useless caracters (encoding + end of line)
            column = column.strip("\ufeff")
            column = column.strip("ï»¿")
            column = column.strip("\n")
            name_column = column.split(sep)
            nb_column = len(name_column) # Number of columns

            # Add into data (dictionary)
            # Each key correspond to a column
            for i in range(nb_column):
                data[name_column[i]] = {}
                n, u = self.__parser_name_unit(name_column[i])
                data[name_column[i]]["name"] = n
                data[name_column[i]]["unit"] = u
                data[name_column[i]]["data"] = []
            
            # Read the rest of the file
            lines = file.readline()
            lines = lines.strip('\n')
            lines_iter = 1
            while lines != "":
                list_lines = lines.split(sep)
                
                nb_elems_line = len(list_lines)
                # Check that the number of elements of a row is equal to the number of columns
                if nb_elems_line != nb_column:
                    raise ValueError(f"La ligne {lines_iter+1} a {nb_elems_line} éléments mais il y a {nb_column} colonnes")
                
                # Add the row elements into data
                # The elements are in a list where the key is the name of the column
                for elem_index in range(len(list_lines)):
                    elem, type_elem = self.__convert_element(list_lines[elem_index])
                    # If it's the first element, we add his type to the dictionary
                    if lines_iter == 1:
                        data[name_column[elem_index]]["type"] = type_elem
                    # Else, we compare if it's the same type
                    else:
                        correct_type = data[name_column[elem_index]]["type"]
                        # If it's an int or a float, we accept both of them
                        if correct_type == int and type_elem == float:
                            correct_type = float
                            data[name_column[elem_index]]["type"] = correct_type
                        elif correct_type == float and type_elem == int:
                            for i in range(len(elem)):
                                elem[i] = float(elem[i])
                            type_elem = float
                        if type_elem != correct_type:
                            raise TypeError(f"L'élément à la ligne {lines_iter+1} et colonne {'.'.join(data[name_column[elem_index]]['name'])} est du type {self.__type_to_str(type_elem)} au lieu du type {self.__type_to_str(correct_type)}")
                    
                    # Add into data
                    data[name_column[elem_index]]["data"].append(elem)
                    
                # Next line
                lines = file.readline()
                lines = lines.strip('\n')
                lines_iter += 1
                
        return data
    

    def __convert_element(self, val_str: str) -> Tuple[List[Union[bool, int, float, str]], type]:
        """ Convert a value in a string to his corresponding type

        Args:
            val_str (str): The value in a string

        Returns:
            type: The type of the value
        """
        constructors = [int, float]
        
        # Change the value to their corresponding type
        val_list = val_str.split(',')
        val_list_save = val_list.copy()
        for i in range(len(val_list)):
            if val_list[i].lower() == "true":
                val_list[i] = True
            elif val_list[i].lower() == "false":
                val_list[i] = False
            else:
                for c in constructors:
                    try:
                        val_list[i] = c(val_list[i])
                        break
                    except ValueError:
                        pass   
        # Check if their are all the same type
        type_val_ref = type(val_list[0])
        for i in range(1, len(val_list)):
            # If there is one who have not the same type, return list of string
            if not isinstance(val_list[i], type_val_ref):
                return val_list_save, str
        return val_list, type_val_ref
            
    
    def __parser_name_unit(self, name_col: str) -> Tuple[List[str], List[str]]:
        """ Parse the 'name' of a column to get the 'original' name and the unit

        Args:
            name_col (str): Format of the column name :
                - nom
                - nom.unite
                - nom.nom.unite
                - nom.unite
                - nom.nom.unite.unite
                etc...

        Returns:
            tuple: Returns a tuple with the name and the unit of 'name_col'.
            For example :
                - 'distance'            -> (['distance'], [])
                - 'distance.km'         -> (['distance'], ['km'])
                - 'vitesse.km.h'        -> (['vitesse'], ['km', 'h'])
                - 'temps.distance.km.h' -> (['temps', 'distance'], ['km', 'h'])
        """
        name_list = name_col.split('.')
        
        # Search the index of the 1st unit
        index_spe_name_unit = -1
        for i in range(len(name_list)):
            if name_list[i] in self.__units:
                index_spe_name_unit = i
                break
        
        # If there are no units, then there is only the name of the column
        if index_spe_name_unit == -1:
            return (name_list, [])
        
        return (name_list[0:index_spe_name_unit], name_list[index_spe_name_unit:len(name_list)])


    def __read_xlsx(self, filename: str, engine: str = 'pandas') -> Dict[str, Dict[str, List[Union[List[Union[int, float, bool, str]], str]]]]:
        """ Read a XLSX file with the engine 'pandas' or 'openpyxl'
        Lecture d'un fichier excel avec comme moteur pandas ou openpyxl

        Args:
            filename (str): Path to file
            engine (str, optional): Reading engine ('pandas', 'openpyxl'). Defaults to 'pandas'.

        Raises:
            ValueError: A reading engine different from 'pandas' and 'openpyxl'

        Returns:
            dict: Dictionary with the data of the file if the reading is correct
        """
        match engine:
            case 'pandas':
                try:
                    return self.__read_xlsx_pandas(filename)
                except:
                    return self.__read_xlsx_openpyxl(filename)
            case 'openpyxl':
                try:
                    return self.__read_xlsx_openpyxl(filename)
                except:
                    return self.__read_xlsx_pandas(filename)
            case _:
                raise ValueError(f"'{engine}' n'est pas un moteur de lecture. Utilisez 'pandas' ou 'openpyxl'")
            
       
    def __read_xlsx_pandas(self, filename: str) -> Dict[str, Dict[str, List[Union[List[Union[int, float, bool, str]], str]]]]:
        """ Read a XLSX file with pandas

        Args:
            filename (str): Path to file

        Raises:
            IOError: A problem occur with read_excel() of pandas

        Returns:
            dict: Dictionary with the data of the file if the reading is correct
        """
        try:
            import pandas
            # Read the file
            data = pandas.read_excel(filename)
            return self.__transform_data_pandas(data.to_dict('split'))       
        except:
            raise IOError("Problème rencontré. Lecture impossible avec 'pandas'. Essayez avec 'openpyxl'")
    
    
    def __read_xlsx_openpyxl(self, filename: str) -> Dict[str, Dict[str, List[Union[List[Union[int, float, bool, str]], str]]]]:
        """ Read a XLSX file with openpyxl

        Args:
            filename (str): Path to file

        Raises:
            IOError: A problem occur with the functions used with openpyxl

        Returns:
            dict: Dictionary with the data of the file if the reading is correct
        """
        try:
            import openpyxl
            # Open the file and activate the sheet
            wb = openpyxl.load_workbook(filename=filename)
            sheet = wb.active
            
            # Get the number of rows and columns
            nb_columns = sheet.max_column
            nb_rows = sheet.max_row
            
            data = {}
            name_column = []
            # Read the 1st line to get the name of columns
            for col in range(1, nb_columns+1):
                # Get the value in the cell of row 1 and column col
                cell = sheet.cell(row=1, column=col)
                data[str(cell.value)] = {}
                n, u = self.__parser_name_unit(str(cell.value))
                data[str(cell.value)]["name"] = n
                data[str(cell.value)]["unit"] = u
                data[str(cell.value)]["data"] = []
                name_column.append(str(cell.value))
                
            # Read data
            for row in range(2, nb_rows+1):
                # Get the value in the cell of row row and column col
                for col in range(1, nb_columns+1):
                    cell = sheet.cell(row=row, column=col)
                    if row == 2:
                        data[name_column[col-1]]["type"] = type(cell.value)
                    data[name_column[col-1]]["data"].append([cell.value])
            return data
        except:
            raise IOError("Problème rencontré. Lecture impossible avec 'openpyxl'. Essayez avec 'pandas'")
        
        
    def __transform_data_pandas(self, data: Dict[str, List[Union[str, bool, int, float]]]) -> Dict[str, Dict[str, List[Union[List[Union[int, float, bool, str]], str]]]]:
        """ Transorm a dictionary from pandas to our corresponding dictionary
        Pandas dictionary:
        {
            'index': [rows name],
            'columns: [columns name],
            'data': [list of elements for each row]
        }

        Args:
            data (dict): Pandas dictionary

        Returns:
            dict: Dictionary with the corresponding format
        """
        columns = data['columns']
        values_columns = data['data']
        data_transform = {}
        for c in range(len(columns)):
            data_transform[str(columns[c])] = {}
            n, u = self.__parser_name_unit(str(columns[c]))
            data_transform[str(columns[c])]["name"] = n
            data_transform[str(columns[c])]["unit"] = u
            data_transform[str(columns[c])]["data"] = []
            for l in range(len(values_columns)):
                val = values_columns[l][c]
                # Add type if it's the first value
                if l == 0:
                    data_transform[str(columns[c])]['type'] = type(val)
                data_transform[str(columns[c])]["data"].append([val])
                
        return data_transform


    def __type_to_str(self, type: type) -> str:
        """ Transform a type to a string

        Args:
            type (type): The type

        Returns:
            str: the type in a string
        """
        return str(type).split("'")[1]
